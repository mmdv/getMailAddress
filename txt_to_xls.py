# 引入openpyxl
import openpyxl as xl
import re

#获取txt文件内容
def get_content_from_txt(txt_file):
    data_list = []
    file_object = open(txt_file,encoding='gb18030',errors='ignore')
    try:
        for line in file_object:
            data_list.append(line)
    finally:
        file_object.close()
    # print(dataList)
    return data_list

#数据处理函数
def operate_data(data):
    # print(data)

    # 按空格切分字符串
    reg = '\S+\s'
    res = re.findall(reg,str(data))
    print(res)
    return res

#写入xls文件
def write_xls(data,xls_file):
    # 创建wb对象
    wb = xl.Workbook()
    # 激活
    ws = wb.active
    # 创建表
    ws.title = "sheet1"

    # print(data[0]) #330 陈东	苏州***有限公司	董事 副总经理	chen.dong@****.com
    for i in data:
        print(i)
        # 数据处理
        items = operate_data(i)
        ws.append(items)#行写入append([1,2,3])


    # 列写入方法
    # for i in range(data.__len__()):
    #     # 用 + 拼接不能使用
    #     ws["A%d" % (x + 1)].value = data[i]
    #
    #     # 写入文件
    #     wb.save(xls_file + '.xlsx')
    #     wb = xl.Workbook()
    #     ws = wb.active
    #     ws.title = "sheet1"

    # 写入表
    wb.save(xls_file  + '.xlsx')

# 主函数
if __name__ == '__main__':
    #txt源文件路径
    txt_file =  'E:\\pppsource\\苏州数据.txt'
    #excel结果路径
    xls_file = 'E:\\pppresult\\苏州数据'
    #读取txt
    txt_data = get_content_from_txt(txt_file)

    # 写入xls
    write_xls(txt_data,xls_file)
