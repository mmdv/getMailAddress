import openpyxl as xl
import os
#openpyxl读
def open():
    wb = xl.load_workbook('E:/pppresult/openpyxl/test.xlsx')
    ws = wb['sheet1']
    for row in ws.iter_cols(min_row=1, max_col=1,max_row=2):
        for cell in row:
            print(cell.value)

#openpyxl写  iter_cols  /  ["A1"]
def write():
    wb = xl.Workbook()
    data = [1,2,3,4,5]
    ws = wb.active
    ws.title = "sheet1"
    # ws['A6'] = '2'
    #追加到一行
    # for row in range(1):
    #     ws.append(data)
    #写入具体单元格
    # ws['A4'] = 4

    #列写入方法1
    # len = data.__len__()
    # i = 0
    # for row in ws.iter_cols(min_row=1, max_col=1, max_row=len):
    #     for cell in row:
    #         cell.value = data[i]
    #         i += 1

    #列写入方法2
    for i in range(data.__len__()):
        #用 + 拼接不能使用
        ws["A%d" % (i+1)].value = data[i]

    # 工作簿保存到磁盘
    wb.save('E:/pppresult/openpyxl/test1.xlsx')
