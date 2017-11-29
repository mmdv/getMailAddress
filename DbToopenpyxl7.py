"""
取出数据库数据写入excel 48000条/表  OPENPYXL
"""

import xlwt
import datetime
import pymysql
import re
import os
import openpyxl as xl

#数据库查询全部邮箱
def selectFromDb():
    global selectDbTable
    db = pymysql.connect(host='localhost', port=3306, user='root', passwd='', db='db3', charset='utf8')
    cursor = db.cursor()
    try:
        sql = "SELECT EMAIL FROM " + selectDbTable
        # 执行SQL语句
        cursor.execute(sql)
        # 获取所有记录列表
        results = cursor.fetchall()
        db.close()
        return results
    except Exception as err:
        print('读取数据库出错',err)

#email地址过滤
def getEmails(arg):
    #将每个数据作为字符串,findall邮箱
    regex = "[a-zA-Z0-9][a-zA-Z0-9.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9.]+"
    email = re.findall(regex,str(arg),re.IGNORECASE)  #返回值为 list
    # while '' in email:
    #     email.remove('')
    return email[0]

#写入excel/
def writeExcel(data,lines):#data可以不需要而直接写入excel提升效率?
    print("写ru数据表被执行",data. __len__())
    global writeCount,savePath
    excelNameCount =  fileCount(savePath) - 1#获取存储文件夹文件数
    print('获取到文件数',excelNameCount)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "sheet1"
    x = 0
    # 列写入方法2
    for i in range(data.__len__()):
        # 用 + 拼接不能使用
        ws["A%d" % (x + 1)].value = data[i]
        #写入列从0开始计数
        x += 1

        if i % lines == 0 and i != 0:#非第一次,不存空表
            # 写入文件
            wb.save(savePath + str(excelNameCount) + '.xlsx')
            x = 0#重置0
            writeCount += 1 #统计次数
            excelNameCount += 1 #文件名次数+
            wb = xl.Workbook()
            ws = wb.active
            ws.title = "sheet1"
    #最后一次写入
    wb.save(savePath + str(excelNameCount) + '.xlsx')
# 获取写入目标文件夹的文件数,用于计数命名
def fileCount(savePath):
    for parent, dirnames, filenames in os.walk(savePath):
        # print('文件数..................', filenames.__len__())
        # 获取当前文件夹写文件数,继续数目新增命名.xls文件
        fileTotal = filenames.__len__()
    return fileTotal

if __name__ == "__main__":
    writeCount = 0
    starttime = datetime.datetime.now()
    #每表数据条数
    lines = 480000
    #查表名
    selectDbTable = " EMAILTEST"
    #存储路径
    # savePath = "E:/pppresult/openpyxl/email/"
    savePath = "E:/pppresult/fromlg/"
    # 清洗结果
    results = []
    #获取邮件地址
    datas = selectFromDb()
    print(datas)
    for email in datas:
        try:
            temp = getEmails(email[0])
            if temp:
                results.append(getEmails(email[0]))
        except:
            print('空数据')

    writeExcel(results,lines)
    endtime = datetime.datetime.now()
    print('运行时间:', endtime - starttime)